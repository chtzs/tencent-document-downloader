from openpyxl import Workbook
import json
from typing import List
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# I know you are confusing with those 'magic number'.
# I'm confused either.
CELL_WIDTH_RATIO = 1 / 8
CELL_HEIGHT_RATIO = 0.74754
CELL_DEFAULT_WIDTH = 107 * CELL_WIDTH_RATIO
CELL_DEFAULT_HEIGHT = 24 * CELL_HEIGHT_RATIO


class Cell:
    def __init__(self, index=0, text="", color="#000000", background_color="#FFFFFF", foreground_color="#000000",
                 font_family="微软雅黑", font_size=11, text_align="left", bold="") -> None:
        self.index = index
        self.text = text
        self.color = color
        self.background_color = background_color
        self.foreground_color = foreground_color
        self.font_family = font_family
        self.font_size = font_size
        self.text_align = text_align
        self.bold = bold


class MergeInfo:
    def __init__(self,  start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        self.start_row = start_row
        self.end_row = end_row
        self.start_col = start_col
        self.end_col = end_col


class SizeInfo:
    def __init__(self, type: str, no: int, size: int) -> None:
        self.type = type
        self.no = no
        self.size = size


def parse_cells(content: List) -> List[Cell]:
    cells = content[1]
    styles = content[2]["0"]
    re = []
    for index, cell in cells.items():
        c = Cell(index=int(index))
        # has text?
        if "2" in cell:
            c.text = cell["2"][1]
        # has style?
        if "8" in cell:
            style = cell["8"][0][1]
            if "0" in style and style["0"] != '':
                c.color = style["0"]
            if "1" in style and style["1"] != '':
                c.font_family = style["1"]
            if "2" in style and style["2"] != '':
                c.font_size = style["2"]
        # has more style?
        if "3" in cell:
            style_index = cell["3"]
            style = styles[style_index]
            if "3" in style:
                if style["3"][0] == "#":
                    c.background_color = style["3"]
            if "8" in style:
                c.text_align = style["8"]
            if "10" in style:
                c.bold = style["10"]
            if "11" in style:
                c.foreground_color = style["11"]

        re.append(c)

    return re


def parse_merge(merge_info: List) -> MergeInfo:
    return MergeInfo(merge_info[1], merge_info[2], merge_info[3], merge_info[4])


def parse_size(size_info: List) -> SizeInfo:
    if size_info[2][0][0] == 2 or size_info[1][0][0] != size_info[1][0][1]:
        return None
    s = SizeInfo(type=size_info[3], no=size_info[1][0][0],
                 size=size_info[2][0][2])
    return s


def parse_image(images: List) -> List[Cell]:
    pass


class SheetGenerator:
    def __init__(self, worksheet: Worksheet, sheet_content: dict, max_col: int) -> None:
        self.worksheet = worksheet
        self.sheet_content = sheet_content
        self.max_col = max_col
        self.cells: List[Cell] = []
        self.merge_infos: List[MergeInfo] = []
        self.size_infos: List[SizeInfo] = []
        self.image_infos = []
        self.parse_data()

    def parse_data(self) -> List[Cell]:
        for attribute in self.sheet_content:
            for e in attribute:
                # Content of cells
                t = e["t"]
                c = e["c"]
                if t == 2:
                    size_info = parse_size(c)
                    if size_info is not None:
                        self.size_infos.append(size_info)
                elif t == 3:
                    self.cells.extend(parse_cells(c))
                # Merge information
                elif t == 5:
                    self.merge_infos.append(parse_merge(c))
                # Image
                elif t == 8:
                    pass
                    # image_infos.append(self.parse_image(e.c))

    def generate_sheet(self) -> Worksheet:
        max_row = 0
        for cell in self.cells:
            row = cell.index // self.max_col + 1
            col = cell.index % self.max_col + 1
            max_row = max(max_row, row)
            c = self.worksheet.cell(row=row, column=col, value=cell.text)
            font = Font(
                name=cell.font_family,
                size=cell.font_size,
                color=cell.color[1:]
            )
            s = Side(style='medium', color=cell.background_color[1:])
            c.border = Border(left=s, right=s, top=s, bottom=s)
            c.font = font
            c.fill = PatternFill(patternType="solid",
                                 fgColor=cell.background_color[1:])
            c.alignment = Alignment(
                horizontal=cell.text_align, vertical="center", wrap_text=True)

        for col in range(1, self.max_col + 1):
            self.worksheet.column_dimensions[get_column_letter(
                col)].width = CELL_DEFAULT_WIDTH

        # for row in range(1, max_row + 1):
        #     self.worksheet.row_dimensions[row].height = CELL_DEFAULT_HEIGHT

        for merge_info in self.merge_infos:
            self.worksheet.merge_cells(
                start_row=merge_info.start_row + 1,
                start_column=merge_info.start_col + 1,
                end_row=merge_info.end_row + 1,
                end_column=merge_info.end_col + 1
            )

        for size_info in self.size_infos:
            if size_info.type == "COLUMNS":
                self.worksheet\
                    .column_dimensions[get_column_letter(size_info.no + 1)]\
                    .width = size_info.size * CELL_WIDTH_RATIO
            elif size_info.type == "ROWS":
                self.worksheet\
                    .row_dimensions[size_info.no + 1]\
                    .height = size_info.size * CELL_HEIGHT_RATIO
            else:
                raise Exception(
                    "Unexpected type of SizeInfo: " + size_info.type)
        return self.worksheet

    def get_sheet_dict(self):
        # 1. 获取单元格信息
        sheet = {}
        for cell in self.cells:
            row_no = cell.index // self.max_col
            col_no = cell.index % self.max_col
            # 创建一个数组来代表"行"
            if not (row_no in sheet):
                row = [None for i in range(0, self.max_col)]
                sheet[row_no] = row
            else:
                row = sheet[row_no]
            # 填充信息
            row[col_no] = cell.text

        # 2. 将合并的单元格拆分
        for merge_info in self.merge_infos:
            text = sheet[merge_info.start_row][merge_info.start_col]
            # if text == '' or text == None:
            #     print("Error!")
            #     exit(0)
            for row in range(merge_info.start_row, merge_info.end_row + 1):
                for col in range(merge_info.start_col, merge_info.end_col + 1):
                    sheet[row][col] = text
        return sheet


if __name__ == '__main__':
    with open("test/BB08J4.json", "r", encoding="utf-8") as f:
        content = f.read()
    data = json.loads(content)
    wb = Workbook()
    ws = wb.create_sheet("Test")
    generator = SheetGenerator(
        ws, data["data"]["initialAttributedText"]["text"][0], 13)
    empty_ws = wb["Sheet"]
    wb.remove(empty_ws)
    wb.save("./test.xlsx")
    print("Done.")
